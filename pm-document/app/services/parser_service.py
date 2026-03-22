import logging
import os

logger = logging.getLogger(__name__)


def extract_text(file_path: str) -> str:
    """파일에서 텍스트 추출. PDF, DOCX, TXT, Excel, HWP 지원."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".pdf":
        return _extract_pdf(file_path)
    elif ext == ".docx":
        return _extract_docx(file_path)
    elif ext in (".txt", ".md", ".csv"):
        return _extract_text(file_path)
    elif ext in (".xlsx", ".xls"):
        return _extract_excel(file_path)
    elif ext == ".hwp":
        return _extract_hwp(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _extract_pdf(file_path: str) -> str:
    import fitz  # PyMuPDF

    doc = fitz.open(file_path)
    text_parts = []
    for page in doc:
        text_parts.append(page.get_text())
    doc.close()
    text = "\n".join(text_parts)
    logger.info("PDF extracted: %d chars from %s", len(text), file_path)
    return text


def _extract_docx(file_path: str) -> str:
    from docx import Document

    doc = Document(file_path)
    text = "\n".join(p.text for p in doc.paragraphs)
    logger.info("DOCX extracted: %d chars from %s", len(text), file_path)
    return text


def _extract_text(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8") as f:
        text = f.read()
    logger.info("Text extracted: %d chars from %s", len(text), file_path)
    return text


def _extract_excel(file_path: str) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    text_parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        text_parts.append(f"[시트: {sheet}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(cells):
                text_parts.append("\t".join(cells))
    wb.close()
    text = "\n".join(text_parts)
    logger.info("Excel extracted: %d chars from %s", len(text), file_path)
    return text


def _extract_hwp(file_path: str) -> str:
    import olefile

    ole = olefile.OleFileIO(file_path)
    text_parts = []

    if ole.exists("PrvText"):
        prv_text = ole.openstream("PrvText").read()
        text_parts.append(prv_text.decode("utf-16-le", errors="ignore"))
    elif ole.exists("BodyText/Section0"):
        # Section 스트림에서 텍스트 추출 (간이 파서)
        for entry in ole.listdir():
            if entry[0] == "BodyText":
                stream = ole.openstream(entry)
                raw = stream.read()
                # HWP 바이너리에서 텍스트 추출 (UTF-16LE)
                decoded = raw.decode("utf-16-le", errors="ignore")
                # 제어 문자 제거
                cleaned = "".join(c for c in decoded if c.isprintable() or c in "\n\t")
                text_parts.append(cleaned)

    ole.close()
    text = "\n".join(text_parts)
    logger.info("HWP extracted: %d chars from %s", len(text), file_path)
    return text
